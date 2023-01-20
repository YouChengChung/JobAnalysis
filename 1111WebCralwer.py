# -*- coding: utf-8 -*-
"""
Created on Dec 2022
@author: Y.C
"""
#aa="面議 經常性48萬/年"
from bs4 import BeautifulSoup
import requests #Method:get or post
#import pandas as pd
import openpyxl
import time

def check_numberdigit(x):
    if len(x)<5:
        return int(float(x)*10000)
    else:
        return int(x)

def adjust_salary(x): #處理薪資
    each_salary=[]
    if ("月" in x) or ("年" in x):
        each_number=""
        for i in x:
            if i.isdigit() or i=="." or i =="~":
                if i=="~":
                    each_salary.append(check_numberdigit(each_number))
                    each_number="" 
                else:
                    each_number+=i
        each_salary.append(check_numberdigit(each_number))       
        if len(each_salary)<2:
            each_salary.append(each_salary[0])      
        if "年" in x:
            each_salary[0],each_salary[1]=int(each_salary[0]/12),int(each_salary[1]/12)      
    else:
        each_salary.append("時薪/論件")
        each_salary.append("時薪/論件")
    return each_salary[0],each_salary[1]

"""用openpyxl寫進worksheet"""
wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = "職缺名稱"
ws['B1'] = "公司名稱"
ws['C1'] = "職缺地區"
ws['D1'] = "薪資待遇1"
ws['E1'] = "職缺連結2"
ws['F1'] = "職缺連結"

x=1
starttime = time.ctime()
for i in range(1,71):
    time.sleep(3)
    URL = "https://www.1111.com.tw/search/job?ks=%E5%A4%A7%E6%95%B8%E6%93%9A&page="+str(i)
    
    res = requests.get(URL)

    soup=BeautifulSoup(res.text,"html.parser")
    total = soup.find_all("div","job_item_info")
    
    for i in total:
        
        print("(",x,")","-"*50,sep="")
        #title = i.text
        title = i.find("h5").text
        company_name=i.find("h6").text
        location = i.find("a","job_item_detail_location mr-3 position-relative").text
        salary = i.find("div","job_item_detail_salary ml-3 font-weight-style digit_6").text
        herf = i.a["href"]
        
        print("職缺：",title)
        print("公司：",company_name)
        print("地點：",location)
        print("薪資：",salary)
        print("連結：",herf)
        
        salary_low,salary_high = adjust_salary(salary)
        x+=1
        #寫進worksheet
        ws.append([title,company_name,location,salary_low,salary_high,herf])

